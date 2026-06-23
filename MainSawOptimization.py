import io
import json
import os
import pandas as pd
import streamlit as st
from collections import defaultdict
from datetime import datetime

# ---------------- Config ----------------
STOCK = 4800.0
MIN_REUSE = 330.0            # strictly > 330 mm is reusable
MEMORY_FILE = "offcut_memory.csv"  # local CSV for persistent offcuts
APP_TITLE = "🪚 Cut Batch Optimizer (with Offcut Memory)"

# Materials to ignore completely (do not optimize, do not output). Exact match after trim.
EXCLUDED_MATERIALS = {
    "WB1183 (S-135-OW)",
    "CV-03 MDF",
    "190772-White",
    "9704.58.CFBAL",
    "CV0430-07-16",
    "WB1112 (S-126 OW)",
    "WB1093 (S-126 Black)",
}

# Main Saw machine materials (exact match)
MAIN_SAW_MATERIAL_CODES = {
    "1749.05.00.4880",
    "1769.05.33.4880",
    "3216.05.33.4880",
    "3223.05.33.4880",
    "3712.05.33.4880",
    "3715.05.33.4880",
    "3832.05.33.4880",
    "3845.05.33.4880",
    "3952.05.33.4880", 
    "3955.05.33.4880",
    "4710.05.00.4880",
    "5841.05.00.4880",
    "5853.05.00.4880",
    "5821.05.00.4880",
    "5852.05.00.4880",
    "5873.05.00.4880",
    "9641 - VYJE - 450",
    "9640 - VYJE - 362",
    "3952.05.00.4880",
    "5856.05.00.4880",
}

# Materials that must be exported to a separate file and optimized last.
# Matching policy:
# - exact match to full material string (after trim)
SAW13_MATERIAL_CODES = {
    "5760.05.00.4880",
    "5762.05.00.4880",
    "5901.05.00.4880",
    "5919.05.33.4880",
    "5956.05.00.4880",
    "5969.05.00.4880",
    "4760.05.00.4880",
    "4762.05.00.4880",
    "927270B (9935)",
    "927270Brz (9935)",
    "927270C (9935)",
    "927270Cus (9935)",
    "927270M (9935)",
    "927270N (9935)",
    "927270S (9935)",
    "927270W (9935)",
    "927827B (9964)",
    "927827Brz (9964)",
    "927827C (9964)",
    "927827Cus (9964)",
    "927827M (9964)",
    "927827N (9964)",
    "927827S (9964)",
    "927827W (9964)",
    "99301-NF",
}

# Tiger Stop Saw machine materials (exact match)
TIGERSTOP_MATERIAL_CODES = {
    "5883.05.00.4880",
    "5825.05.00.4880",
    "5772.05.00.4880",
    "5872.05.00.4880",
    "1800.05.00.4880",
    "5771.05.00.4880",
    "5890.05.00.4880",
    "5800.05.00.4880",
    "5826.05.00.4880",
    "5881.05.00.4880",
    "5828.05.00.4880",
    "5801.05.00.4880",
    "5803.05.00.4880",
    "5959.05.00.4880",
    "5770.05.00.4880",
    "5928.05.00.4880",
    "5832.05.00.4880",
    "5846.05.00.4880",
    "5903.05.00.4880",
    "5915.05.00.4880",
    "5856.05.00.4880",
    "5960.05.00.4880",
    "5967.05.00.4880",
    "4761.05.00.4880",
    "1880.05.00.4880",
    "1809.05.00.4880",
    "1830.05.00.4880",
    "1831.05.00.4880",
    "1821.05.00.4880",
    "8852.00.00.4880",
    "9847.00.00.4880",
}

# Saw #13 JSON export: material-specific MaxY map
MAXY_MAP = {
    "5760.05.00.4880": 45.0,
    "5762.05.00.4880": 45.0,
    "5956.05.00.4880": 9.0,
    "5919.05.33.4880": 25.0,
    "5901.05.00.4880": 45.3,
    "R8572FS-836": 45.0,
    "3223.05.33.4880": 45.0,
    "5969.05.00.4880": 45.0,
    "4760.05.00.4880": 45.0,
    "4762.05.00.4880": 45.0,
    "927270B (9935)": 45.0,
    "927270Brz (9935)": 45.0,
    "927270C (9935)": 45.0,
    "927270Cus (9935)": 45.0,
    "927270M (9935)": 45.0,
    "927270N (9935)": 45.0,
    "927270S (9935)": 45.0,
    "927270W (9935)": 45.0,
    "927827B (9964)": 45.0,
    "927827Brz (9964)": 45.0,
    "927827C (9964)": 45.0,
    "927827Cus (9964)": 45.0,
    "927827M (9964)": 45.0,
    "927827N (9964)": 45.0,
    "927827S (9964)": 45.0,
    "927827W (9964)": 45.0,
    "99301-NF": 45.0,
}

DEFAULT_MAXY = 45.0
MAXZ = 83.1
BAR_TYPE = 2
BAR_ID_START = 134
FIXED_DXF_NAME = "SL-LHB8004-20T"
MIN_SAW13_CUT_LENGTH = 570.0
SAW13_CUT_WASTE = 76.2
SAW13_JSON_REQUIRED_COLUMNS = [
    "qty",
    "length",
    "inangle",
    "outangle",
    "Job",
    "itemId",
    "material",
    "bin",
    "optimized_group",
    "source_length_mm",
]
BARCAD_URL = (
    "http://wincc.oss-cn-qingdao.aliyuncs.com/"
    "Wincc_dandongchuangyanjian/material/profile/image/"
    "barCad_182c74ad-02f4-4ba2-9de1-8171424f9168.png"
)
BARCAD_FILE_URL = (
    "http://wincc.oss-cn-qingdao.aliyuncs.com/"
    "Wincc_dandongchuangyanjian/material/profile/image/"
    "barCad_22bbbd92-13c6-4baf-ab21-b5d1051b161c.dxf"
)

# Saw #13 qty handling mode:
# - "multiply_length": each row consumes (length * qty) in optimization.
# - "ignore_qty": legacy behavior, optimization uses row length only.
SAW13_QTY_OPTIMIZATION_MODE = "multiply_length"


def _clean_number(value, decimals=3):
    if pd.isna(value):
        return None
    try:
        value = float(value)
        if value.is_integer():
            return int(value)
        return round(value, decimals)
    except Exception:
        return value


def _clean_text(value):
    if pd.isna(value):
        return ""
    value = str(value).strip()
    if value.endswith(".0"):
        value = value[:-2]
    return value


def _round_half_up_whole_mm(value):
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    try:
        from decimal import Decimal, ROUND_HALF_UP

        return float(Decimal(str(value)).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    except Exception:
        try:
            return float(round(float(value)))
        except Exception:
            return value


def _sort_key(value):
    try:
        return float(value)
    except Exception:
        return str(value)


def _validate_saw13_json_columns(df: pd.DataFrame) -> list[str]:
    return [col for col in SAW13_JSON_REQUIRED_COLUMNS if col not in df.columns]


def generate_saw13_json_from_optimized(df: pd.DataFrame) -> tuple[dict, dict]:
    """Generate Saw #13 JSON from optimized output rows.

    Uses source_length_mm from optimized output as each bar length.
    """
    data = df.copy().dropna(how="all")

    missing_cols = _validate_saw13_json_columns(data)
    if missing_cols:
        raise ValueError("Missing required columns for Saw #13 JSON: " + ", ".join(missing_cols))

    groups = defaultdict(list)
    for _, row in data.iterrows():
        grp = row.get("optimized_group")
        if pd.isna(grp):
            continue
        groups[grp].append(row)

    bars = []
    over_length_groups = []
    missing_maxy_materials = []
    bar_number = 1

    for optimized_group in sorted(groups.keys(), key=_sort_key):
        group_rows = groups[optimized_group]
        first_row = group_rows[0]

        material = _clean_text(first_row.get("material"))
        display_group = _clean_text(optimized_group)

        maxy = MAXY_MAP.get(material, DEFAULT_MAXY)
        if material not in MAXY_MAP and material not in missing_maxy_materials:
            missing_maxy_materials.append(material)

        source_length = _round_half_up_whole_mm(first_row.get("source_length_mm"))
        if source_length is None:
            raise ValueError(f"source_length_mm is missing for optimized_group {display_group}")

        bar_length = float(source_length)

        pieces = []
        piece_number = 1
        total_cut_length_expanded = 0.0

        for row_index, row in enumerate(group_rows, start=1):
            qty_value = row.get("qty", 1)
            qty = int(float(qty_value)) if not pd.isna(qty_value) else 1

            item_id = _clean_text(row.get("itemId"))
            bin_name = _clean_text(row.get("bin"))
            job = _clean_text(row.get("Job"))

            length = _round_half_up_whole_mm(row.get("length"))
            left_angle = _round_half_up_whole_mm(row.get("inangle"))
            right_angle = _round_half_up_whole_mm(row.get("outangle"))

            for qty_index in range(1, qty + 1):
                total_cut_length_expanded += float((length or 0) + SAW13_CUT_WASTE)

                pieces.append({
                    "PieceNumber": piece_number,
                    "Num": 1,
                    "Over_status": 0,
                    "BarCode": (
                        f"{item_id}-{bin_name}-G{display_group}-"
                        f"R{row_index:02d}-Q{qty_index:02d}"
                    ),
                    "Customer": f"JOB {job}",
                    "Cc_code": job,
                    "Buyer": item_id,
                    "BuyAddress": bin_name,
                    "CtrsX": "0.0",
                    "CtrsYp": "0.0",
                    "CtrsYm": "0.0",
                    "Mirror": "0.0",
                    "Extra_Pie": None,
                    "OrderCode": item_id,
                    "GroupCode": f"G{display_group}",
                    "Cut": {
                        "Length": length,
                        "sxB": left_angle,
                        "dxB": right_angle,
                        "sxC": 90,
                        "dxC": 90,
                    },
                    "Work": [],
                    "Machining": [],
                })
                piece_number += 1

        calculated_left = round(bar_length - total_cut_length_expanded, 3)
        bar_left_length = max(0, calculated_left)

        if calculated_left < 0:
            over_length_groups.append({
                "optimized_group": display_group,
                "group_code": f"G{display_group}",
                "material": material,
                "bar_length": bar_length,
                "total_cut_length_expanded": round(total_cut_length_expanded, 3),
                "calculated_bar_left_length": calculated_left,
                "written_bar_left_length": bar_left_length,
            })

        bars.append({
            "DXF_Name": FIXED_DXF_NAME,
            "BarNumber": bar_number,
            "MaxY": maxy,
            "MaxZ": MAXZ,
            "Length": bar_length,
            "BarName": f"G{display_group} - {material}",
            "BarColor": "/",
            "BarLeftLength": bar_left_length,
            "BarType": BAR_TYPE,
            "BarCad": BARCAD_URL,
            "BarId": BAR_ID_START + bar_number - 1,
            "BarCad_file": BARCAD_FILE_URL,
            "SortingCode": f"G{display_group}",
            "Machining": {"Faces": []},
            "profile_type": None,
            "profile_type_sort": 1,
            "Piece": pieces,
        })
        bar_number += 1

    json_data = {
        "code": 200,
        "message": "操作成功!",
        "data": {
            "Version": "PS-20210825-16",
            "Generator": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Unit": None,
            "Name": "SAW 13 - OPTIMIZED EXPORT - SOURCE LENGTH - QTY EXPANDED",
            "Bars": bars,
        },
        "meta": None,
    }

    summary = {
        "excel_rows_read": int(len(data)),
        "bars_created": len(bars),
        "pieces_created": sum(len(bar["Piece"]) for bar in bars),
        "group_rule": "same as optimized_group",
        "length_rule": "Length comes from source_length_mm per optimized group",
        "maxy_rule": "material-based MaxY map (default 45.0 if missing)",
        "qty_rule": "qty expanded into separate Piece records",
        "cut_waste_mm": SAW13_CUT_WASTE,
        "minimum_cut_length_mm": MIN_SAW13_CUT_LENGTH,
        "over_length_groups": over_length_groups,
        "missing_maxy_materials_defaulted_to_45": missing_maxy_materials,
    }

    return json_data, summary


def is_saw13_material(material_value) -> bool:
    if material_value is None or (isinstance(material_value, float) and pd.isna(material_value)):
        return False
    s = str(material_value).strip()
    if not s:
        return False

    return s in SAW13_MATERIAL_CODES


def is_tigerstop_material(material_value) -> bool:
    """Tiger Stop is the catch-all: anything not Main Saw, not Saw #13, and not Excluded."""
    if material_value is None or (isinstance(material_value, float) and pd.isna(material_value)):
        return False
    s = str(material_value).strip()
    if not s:
        return False
    if s in EXCLUDED_MATERIALS:
        return False
    if s in MAIN_SAW_MATERIAL_CODES:
        return False
    if s in SAW13_MATERIAL_CODES:
        return False
    return True


def is_main_saw_material(material_value) -> bool:
    if material_value is None or (isinstance(material_value, float) and pd.isna(material_value)):
        return False
    s = str(material_value).strip()
    if not s:
        return False
    return s in MAIN_SAW_MATERIAL_CODES


def is_excluded_material(material_value) -> bool:
    if material_value is None or (isinstance(material_value, float) and pd.isna(material_value)):
        return False
    s = str(material_value).strip()
    return s in EXCLUDED_MATERIALS


def is_recognized_material(material_value) -> bool:
    """Return True if the material is not excluded (Tiger Stop catches everything else)."""
    if is_excluded_material(material_value):
        return False
    return True
# ---------------------------------------

st.set_page_config(page_title="Cut Batch Optimizer", page_icon="🪚", layout="centered")
st.title(APP_TITLE)
st.caption("Upload XLS/XLSX/CSV. Optimizes groups with offcut reuse and remembers reusable offcuts for future batches.")

# ---------------------
# Machine-readable format converter
# ---------------------
def transform_optimized_to_machine_readable(optimized_df: pd.DataFrame) -> pd.DataFrame:
    """Convert the optimized output DataFrame into the machine-readable profile CSV layout.

    Notes (as requested):
    - Uses `optimized_group` as the group/page column.
    - Uses `itemId` as the barcode/item id column.
    - Reads from the optimized sheet/output (NOT the original uploaded batch).
    """

    df = optimized_df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    # Trim string cells
    df = df.map(lambda x: x.strip() if isinstance(x, str) else x)

    def _round_half_up_whole_number(value, default: str = "0") -> str:
        """Round to nearest whole number using HALF-UP.

        Examples:
        - 1044.7 -> 1045
        - 1044.3 -> 1044
        - 1044.5 -> 1045
        """
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return default
        s = str(value).strip()
        if s == "":
            return default
        try:
            from decimal import Decimal, ROUND_HALF_UP

            # Use string -> Decimal to avoid binary float rounding surprises.
            d = Decimal(s)
            return str(int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP)))
        except Exception:
            # If it can't be parsed, keep original (but still return a string).
            return s

    def find_col(*names: str) -> str | None:
        lowered = {str(c).strip().lower(): c for c in df.columns}
        for name in names:
            key = str(name).strip().lower()
            if key in lowered:
                return lowered[key]
        return None

    col_group = find_col("optimized_group")
    col_material = find_col("material")
    col_length = find_col("length")
    col_qty = find_col("qty", "quantity")
    col_itemid = find_col("itemId", "itemid", "item_id", "item id")

    if not col_group:
        raise ValueError("Optimized output is missing required column: optimized_group")
    if not col_material:
        raise ValueError("Optimized output is missing required column: material")
    if not col_length:
        raise ValueError("Optimized output is missing required column: length")
    if not col_itemid:
        raise ValueError("Optimized output is missing required column: itemId")

    # If qty is missing, treat each row as 1
    if not col_qty:
        df["_qty"] = "1"
        col_qty = "_qty"

    # Preserve group order as first occurrence
    first_rows = df.groupby(col_group, sort=False).head(1).reset_index(drop=True)
    group_order = list(first_rows[col_group].astype(str))

    # Export ALL groups (no special stop condition)
    kept_groups = group_order

    group_rows = {str(g): df[df[col_group].astype(str) == str(g)].reset_index(drop=True) for g in kept_groups}
    num_pages = len(kept_groups)
    max_items = max((len(gdf) for gdf in group_rows.values()), default=0)

    def get_val(g: str, idx: int, col: str | None, default: str = "") -> str:
        gdf = group_rows[str(g)]
        if idx < len(gdf) and col and col in gdf.columns:
            v = gdf.iloc[idx][col]
            return "" if pd.isna(v) else str(v)
        return default

    rows: list[list[str]] = []
    rows.append(["List separator=", "Decimal symbol=."] + [""] * max(0, num_pages - 2))
    rows.append(["Scheme Scheme"] + [""] * num_pages)
    page_labels = [f"Page_{i+1}" for i in range(num_pages)]
    rows.append(["LANGID_804"] + page_labels)
    rows.append(["LANGID_404"] + page_labels)
    rows.append(["1"] + [str(i + 1) for i in range(num_pages)])

    rows.append(["204_HMI_Scheme_ProjectData_BarchCode"] + ["1"] * num_pages)
    rows.append(["204_HMI_Scheme_ProjectData_EngInfo"] + ["1"] * num_pages)
    rows.append(["204_HMI_Scheme_ProjectData_ProfileName"] + [get_val(g, 0, col_material, "") for g in kept_groups])
    rows.append(["204_HMI_Scheme_ProjectData_ProfileCode"] + ["0"] * num_pages)
    rows.append(["204_HMI_Scheme_ProjectData_RawLength"] + ["4800"] * num_pages)
    rows.append(["204_HMI_Scheme_ProjectData_RawHeight"] + ["0"] * num_pages)
    rows.append(["204_HMI_Scheme_ProjectData_RawWidth"] + ["0"] * num_pages)
    rows.append(["204_HMI_Scheme_ProjectData_Amount"] + ["0"] * num_pages)

    for k in range(1, max_items + 1):
        idx = k - 1
        rows.append([f"204_HMI_Scheme_ProjectData_PerformData{{{k}}}.length"]
                    + [_round_half_up_whole_number(get_val(g, idx, col_length, "0"), default="0") for g in kept_groups])
        rows.append([f"204_HMI_Scheme_ProjectData_PerformData{{{k}}}.angle2"] + ["0"] * num_pages)
        rows.append([f"204_HMI_Scheme_ProjectData_PerformData{{{k}}}.angle1"] + ["0"] * num_pages)
        rows.append([f"204_HMI_Scheme_ProjectData_PerformData{{{k}}}.quantity"]
                    + ["1"] * num_pages)

    for k in range(1, max_items + 1):
        idx = k - 1
        rows.append([f"204_HMI_Scheme_ProjectData_PerformData{{{k}}}.barcode"]
                    + [get_val(g, idx, col_itemid, "") for g in kept_groups])

    return pd.DataFrame(rows).fillna("")


def _extract_material_prefix(material_value) -> str:
    if material_value is None or (isinstance(material_value, float) and pd.isna(material_value)):
        return ""
    s = str(material_value).strip()
    if not s:
        return ""
    return s.split(".")[0].strip()


def reorder_barcode_pdf_to_optimized(pdf_bytes: bytes, optimized_out: pd.DataFrame) -> tuple[io.BytesIO, pd.DataFrame]:
    """Reorder barcode PDF pages to match optimized cut order.

    - Groups PDF pages by material code prefix (e.g. P3223... -> "3223").
    - For each row in original order, pops the first available page with a matching prefix.
    - This is robust to interleaved materials (main saw / tigerstop / saw13) — no fragile
      single-pointer scan that breaks on mismatches.
    - Final output is reordered to match the optimized order (optimized_group, then orig_index).
    """

    try:
        import fitz  # PyMuPDF
    except Exception as e:
        raise RuntimeError(
            "Missing dependency for PDF processing. Install 'pymupdf' (fitz) in your environment."
        ) from e

    import re
    from collections import defaultdict

    src_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pages = []
    # Capture first material code per page from text: P<digits>...
    pattern = re.compile(r"\bP(\d{2,})[A-Za-z0-9]*\b")
    for i in range(src_doc.page_count):
        text = (src_doc.load_page(i).get_text("text") or "")
        m = pattern.search(text)
        code = m.group(1) if m else None
        token = m.group(0) if m else None
        pages.append({"page_index": i, "code": code, "token": token})

    if optimized_out is None or optimized_out.empty:
        raise ValueError("Optimized output is empty; cannot reorder barcodes")
    if "orig_index" not in optimized_out.columns:
        raise ValueError("Optimized output is missing orig_index; cannot reorder barcodes")
    if "material" not in optimized_out.columns:
        raise ValueError("Optimized output is missing material; cannot reorder barcodes")

    # Reconstruct original order from orig_index
    original_order = optimized_out.sort_values(["orig_index"]).reset_index(drop=True)
    original_order["_mat_prefix"] = original_order["material"].apply(_extract_material_prefix)

    # Build per-prefix buckets of page indices (in order of appearance in the PDF).
    # This avoids the fragile single-pointer sequential scan: with multiple interleaved
    # materials the old scan would advance p past pages belonging to later rows whenever
    # a prefix mismatch occurred, permanently losing those pages.
    code_to_pages: dict[str, list[int]] = defaultdict(list)
    skipped_pages: list[int] = []
    for pg in pages:
        if pg["code"] is not None:
            code_to_pages[pg["code"]].append(pg["page_index"])
        else:
            skipped_pages.append(pg["page_index"])

    # Align: for each row (in original order), pop the first page whose prefix matches.
    row_to_page: dict[int, int] = {}
    used_pages: set[int] = set()
    for _, row in original_order.iterrows():
        oid = int(row["orig_index"]) if pd.notna(row["orig_index"]) else None
        if oid is None:
            continue
        want = str(row.get("_mat_prefix", "") or "").strip()
        if not want:
            continue
        avail = code_to_pages.get(want, [])
        if avail:
            page_idx = avail.pop(0)
            row_to_page[oid] = page_idx
            used_pages.add(page_idx)

    # Build reorder list in optimized order
    optimized_order = optimized_out.sort_values(["optimized_group", "orig_index"]).reset_index(drop=True)
    reordered_page_indices: list[int] = []
    missing_rows: list[dict] = []
    for _, row in optimized_order.iterrows():
        oid = int(row["orig_index"]) if pd.notna(row["orig_index"]) else None
        if oid is None:
            continue
        page_idx = row_to_page.get(oid)
        if page_idx is None:
            missing_rows.append({
                "orig_index": oid,
                "optimized_group": row.get("optimized_group"),
                "material": row.get("material"),
            })
            continue
        reordered_page_indices.append(page_idx)

    # Append any remaining pages not used (keeps information, but after reordered list)
    remaining_pages = [p["page_index"] for p in pages if p["page_index"] not in used_pages]
    final_pages = reordered_page_indices + remaining_pages

    # Create new PDF
    new_doc = fitz.open()
    for page_idx in final_pages:
        new_doc.insert_pdf(src_doc, from_page=page_idx, to_page=page_idx)

    out_buf = io.BytesIO(new_doc.tobytes())
    out_buf.seek(0)

    report = pd.DataFrame({
        "metric": [
            "pdf_pages_total",
            "pdf_pages_with_code",
            "pdf_pages_skipped_no_code",
            "rows_total",
            "rows_mapped_to_pdf",
            "rows_missing_pdf",
        ],
        "value": [
            len(pages),
            sum(1 for x in pages if x["code"] is not None),
            len(skipped_pages),
            len(original_order),
            len(row_to_page),
            len(missing_rows),
        ],
    })

    return out_buf, report

def split_barcode_pdf_by_machine(pdf_bytes: bytes) -> dict[str, io.BytesIO]:
    """Split barcode PDF pages into 3 PDFs based on material code on each label.

    Reads PXX... token from each page, takes the digits (e.g. P5760 -> 5760),
    and matches against the first 4 digits of each machine's material list.
    Returns dict with keys 'main_saw', 'saw13', 'tigerstop' -> BytesIO PDFs.
    Pages that don't match any list are skipped.
    """
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        raise RuntimeError("Missing pymupdf for PDF splitting.") from e

    import re

    # Build prefix lookup: first 4 digits of each material code -> machine name
    prefix_to_machine: dict[str, str] = {}
    for code in MAIN_SAW_MATERIAL_CODES:
        prefix_to_machine[code.split(".")[0].strip()] = "main_saw"
    for code in SAW13_MATERIAL_CODES:
        prefix_to_machine[code.split(".")[0].strip()] = "saw13"
    for code in TIGERSTOP_MATERIAL_CODES:
        prefix_to_machine[code.split(".")[0].strip()] = "tigerstop"

    src_doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    pattern = re.compile(r"\bP(\d{2,})[A-Za-z0-9]*\b")

    buckets: dict[str, list[int]] = {"main_saw": [], "saw13": [], "tigerstop": []}

    for i in range(src_doc.page_count):
        text = (src_doc.load_page(i).get_text("text") or "")
        m = pattern.search(text)
        if not m:
            continue
        code = m.group(1)  # e.g. "5760"
        machine = prefix_to_machine.get(code)
        if machine:
            buckets[machine].append(i)

    result: dict[str, io.BytesIO] = {}
    for key in ("main_saw", "saw13", "tigerstop"):
        if buckets[key]:
            new_doc = fitz.open()
            for page_idx in buckets[key]:
                new_doc.insert_pdf(src_doc, from_page=page_idx, to_page=page_idx)
            buf = io.BytesIO(new_doc.tobytes())
            buf.seek(0)
            result[key] = buf

    return result


# --------- Sidebar: Offcut Controls ---------
st.sidebar.header("🧠 Offcut Settings")

# Global precut (applies to all offcuts)
precut_mm = st.sidebar.number_input("Global precut (mm)", min_value=0.0, value=0.0, step=1.0, help="Applied to every offcut; you can change this later.")

# Upload offcut inventory
st.sidebar.subheader("📤 Upload Reusable Offcuts (Optional)")
uploaded_offcuts = st.sidebar.file_uploader(
    "Upload offcut inventory CSV",
    type=["csv"],
    help="Upload a CSV with columns: material, offcut_length, precut_mm"
)

if uploaded_offcuts is not None:
    try:
        offcut_preview = pd.read_csv(uploaded_offcuts)
        uploaded_offcuts.seek(0)  # Reset file pointer
        st.sidebar.success(f"✅ {len(offcut_preview)} offcuts loaded")
        st.sidebar.dataframe(
            offcut_preview[["material", "offcut_length", "precut_mm"]],
            use_container_width=True,
            height=240
        )
        st.sidebar.caption("Scroll the table to see all loaded offcuts.")
    except Exception as e:
        st.sidebar.error(f"❌ Error reading file: {e}")

# --------- File Upload ----------
uploaded = st.file_uploader("Upload your cut batch file (.xls, .xlsx, .csv)", type=["xls", "xlsx", "csv"])

# Barcode PDF upload (for reordering to match optimized output)
barcode_pdf = st.file_uploader("Upload barcode PDF (optional)", type=["pdf"])

def optimize_one_material(subdf: pd.DataFrame, mem_pool: list[float], length_col: str = "length") -> tuple[pd.DataFrame, list[float]]:
    """
    subdf: rows of one material (with columns: orig_index, length, material)
    mem_pool: list of offcut lengths (>330) for THIS material (already precut-adjusted if needed)
    Returns:
      - assignment df for this material
      - updated mem_pool after consumption/creation
    Logic:
      - Reorder rows to minimize waste (descending length).
      - Each group uses one capacity: either the largest usable offcut (>= largest remaining item), else fresh 4880.
      - sum(length) <= capacity (no kerf).
      - Leftover > 330 saved back to mem_pool (carried forward for future batches).
    """
    # items: (orig_index, length), sort desc
    items = (
        subdf[["orig_index", length_col]]
        .sort_values(length_col, ascending=False)
        .values
        .tolist()
    )

    # work on a local copy of pool
    offcuts = sorted([oc for oc in mem_pool if oc > MIN_REUSE], reverse=True)
    groups = []
    mgroup = 1

    while items:
        largest_len = max(L for _, L in items)
        # choose the largest offcut that can fit the largest remaining item
        usable = [oc for oc in offcuts if oc >= largest_len]
        if usable:
            cap = usable[0]
            offcuts.remove(cap)
        else:
            cap = STOCK

        total = 0.0
        placed_idx = []
        for i, (oid, L) in enumerate(items):
            if total + L <= cap:
                placed_idx.append(i)
                total += L

        if not placed_idx:
            placed_idx = [0]
            total = items[0][1]

        leftover = cap - total
        if leftover > MIN_REUSE:
            offcuts.append(leftover)

        groups.append({
            "mat_group": mgroup,
            "capacity": cap,
            "leftover": leftover,
            "ids": [items[i][0] for i in placed_idx]
        })

        placed_set = set(placed_idx)
        items = [items[i] for i in range(len(items)) if i not in placed_set]
        mgroup += 1

    # explode to rows
    rows = []
    for g in groups:
        for oid in g["ids"]:
            rows.append({
                "orig_index": oid,
                "mat_group": g["mat_group"],
                "source_length_mm": g["capacity"],
                "group_wastage_mm": round(g["leftover"], 3),
                "carryover_offcut_mm": round(g["leftover"], 3) if g["leftover"] > MIN_REUSE else None
            })

    # return assignments + updated pool
    return pd.DataFrame(rows), offcuts


def _normalize_offcut_inventory(offcut_df: pd.DataFrame | None, precut_default: float) -> pd.DataFrame:
    """Normalize uploaded offcut inventory to columns: material, offcut_length, precut_mm.

    - Accepts flexible column naming/casing.
    - Drops rows with non-numeric offcut_length.
    - Fills missing precut_mm with precut_default.
    """
    if offcut_df is None or getattr(offcut_df, "empty", True):
        return pd.DataFrame(columns=["material", "offcut_length", "precut_mm"])

    tmp = offcut_df.copy()
    tmp.columns = [str(c).strip() for c in tmp.columns]
    lowered = {str(c).strip().lower(): c for c in tmp.columns}

    def pick_col(*names: str) -> str | None:
        for n in names:
            key = str(n).strip().lower()
            if key in lowered:
                return lowered[key]
        return None

    col_material = pick_col("material")
    col_length = pick_col("offcut_length", "offcutlength", "offcut", "length")
    col_precut = pick_col("precut_mm", "precutmm", "precut")

    if not col_material or not col_length:
        return pd.DataFrame(columns=["material", "offcut_length", "precut_mm"])

    keep_cols = [col_material, col_length] + ([col_precut] if col_precut else [])
    norm = tmp[keep_cols].rename(columns={
        col_material: "material",
        col_length: "offcut_length",
        (col_precut or ""): "precut_mm",
    })

    norm["material"] = norm["material"].astype(str).str.strip()
    norm["offcut_length"] = pd.to_numeric(norm["offcut_length"], errors="coerce")
    if "precut_mm" not in norm.columns:
        norm["precut_mm"] = precut_default
    else:
        norm["precut_mm"] = pd.to_numeric(norm["precut_mm"], errors="coerce").fillna(precut_default)

    norm = norm[norm["offcut_length"].notna()].copy()
    return norm

def optimize_with_memory(df_in: pd.DataFrame, offcut_mem_df: pd.DataFrame, precut_mm: float):
    """
    df_in: uploaded batch (any columns; needs at least 'length' and optional 'material')
    offcut_mem_df: global memory of offcuts across batches
    precut_mm: global adjustment applied to new offcuts saved now (not subtracted from capacity for usage yet).
               (You said keep 0 for now; we store it with each offcut for future use.)
    """
    # Prepare input
    df = df_in.copy()
    df.columns = [str(c).strip() for c in df.columns]
    df = df.reset_index().rename(columns={"index": "orig_index"})
    if "material" not in df.columns:
        df["material"] = "MATERIAL_1"
    df["material"] = df["material"].fillna("MATERIAL_1")

    # Remove excluded materials entirely (do not output, do not optimize).
    df = df.loc[~df["material"].apply(is_excluded_material)].copy()

    # Also remove materials not in any recognized machine list.
    df = df.loc[df["material"].apply(is_recognized_material)].copy()

    df["length"] = pd.to_numeric(df["length"], errors="coerce")
    valid = df[df["length"].notna()].copy()

    # Default optimization length for all machines uses row length.
    valid["optimization_length"] = valid["length"]

    # Saw #13 only: account qty in optimization so source length planning matches
    # one-piece-at-a-time cutting behavior.
    if SAW13_QTY_OPTIMIZATION_MODE == "multiply_length":
        qty_col = next((c for c in valid.columns if str(c).strip().lower() in ("qty", "quantity")), None)
        if qty_col is not None:
            qty_numeric = pd.to_numeric(valid[qty_col], errors="coerce").fillna(1)
            qty_numeric = qty_numeric.clip(lower=1)
            saw13_mask_valid = valid["material"].apply(is_saw13_material)
            # Each cut on Saw #13 consumes (length + SAW13_CUT_WASTE) per piece, expanded by qty.
            # This must match the JSON total so the optimizer plans enough source material.
            valid.loc[saw13_mask_valid, "optimization_length"] = (
                (valid.loc[saw13_mask_valid, "length"] + SAW13_CUT_WASTE)
                * qty_numeric.loc[saw13_mask_valid]
            )

    # Normalize offcut memory/inventory input
    offcut_mem_df = _normalize_offcut_inventory(offcut_mem_df, precut_default=precut_mm)

    # If nothing remains after filtering, return empty outputs but keep offcut carryover logic.
    if valid.empty:
        # Create empty optimization assignment frame so merge adds expected columns.
        opt = pd.DataFrame(columns=[
            "orig_index",
            "material",
            "source_length_mm",
            "group_wastage_mm",
            "carryover_offcut_mm",
            "optimized_group",
        ])

        out = df.merge(opt, on=["orig_index", "material"], how="left")
        out = out.sort_values(["optimized_group", "orig_index"], na_position="last").reset_index(drop=True)

        export_cols_to_drop = {"orig_index", "group"}
        unnamed_cols = {c for c in out.columns if str(c).strip().lower().startswith("unnamed")}
        out_export = out.drop(columns=[c for c in out.columns if c in export_cols_to_drop or c in unnamed_cols], errors="ignore")

        check_main = pd.DataFrame(columns=["optimized_group", "capacity", "sum_lengths", "material", "materials", "ok_capacity", "ok_single_material"])
        check_tigerstop = check_main.copy()
        check_saw13 = check_main.copy()

        offcut_mem_df = _normalize_offcut_inventory(offcut_mem_df, precut_default=precut_mm)
        batch_id = datetime.now().strftime("%Y%m%d-%H%M%S")
        timestamp = datetime.now().isoformat(timespec="seconds")

        new_mem_records = []
        if not offcut_mem_df.empty:
            carry = offcut_mem_df.loc[offcut_mem_df["offcut_length"] > MIN_REUSE].copy()
            for _, r in carry.iterrows():
                new_mem_records.append({
                    "material": r["material"],
                    "offcut_length": float(r["offcut_length"]),
                    "precut_mm": float(r.get("precut_mm", precut_mm)),
                    "batch_id": batch_id,
                    "timestamp": timestamp,
                })
        new_mem_df = pd.DataFrame(new_mem_records, columns=["material", "offcut_length", "precut_mm", "batch_id", "timestamp"])

        main_excel_buf = io.BytesIO()
        with pd.ExcelWriter(main_excel_buf, engine="openpyxl") as writer:
            out_export.to_excel(writer, sheet_name="main saw", index=False)
            check_main.to_excel(writer, sheet_name="Checks", index=False)
        main_excel_buf.seek(0)

        tigerstop_excel_buf = io.BytesIO()
        with pd.ExcelWriter(tigerstop_excel_buf, engine="openpyxl") as writer:
            out_export.iloc[0:0].to_excel(writer, sheet_name="Tiger Stop Saw", index=False)
            check_tigerstop.to_excel(writer, sheet_name="Checks", index=False)
        tigerstop_excel_buf.seek(0)

        saw13_excel_buf = io.BytesIO()
        with pd.ExcelWriter(saw13_excel_buf, engine="openpyxl") as writer:
            out_export.iloc[0:0].to_excel(writer, sheet_name="saw #13", index=False)
            check_saw13.to_excel(writer, sheet_name="Checks", index=False)
        saw13_excel_buf.seek(0)

        return out, check_main, check_tigerstop, check_saw13, main_excel_buf, tigerstop_excel_buf, saw13_excel_buf, new_mem_df

    # Build material order (only materials in this batch).
    # IMPORTANT: optimize special machines at the end to avoid interleaving groups.
    materials = list(valid["material"].astype(str).drop_duplicates())
    materials_main = [m for m in materials if is_main_saw_material(m)]
    materials_tigerstop = [m for m in materials if is_tigerstop_material(m)]
    materials_saw13 = [m for m in materials if is_saw13_material(m)]
    # Keep Saw #13 last (as requested earlier), TigerStop just before it.
    materials_ordered = materials_main + materials_tigerstop + materials_saw13

    # Build a dict of offcut pools per material from memory
    mem_pools = {m: [] for m in materials_ordered}
    if not offcut_mem_df.empty:
        for m in materials_ordered:
            pool = offcut_mem_df.loc[offcut_mem_df["material"].astype(str) == str(m), "offcut_length"].tolist()
            # (precut policy: we are *storing* precut, not subtracting from capacity now)
            mem_pools[m] = pool

    # Optimize per material using its pool
    parts = []
    updated_pools = {}
    next_global_group = 1
    for m in materials_ordered:
        sub = valid[valid["material"] == m]
        length_col = "optimization_length" if is_saw13_material(m) else "length"
        assign_m, pool_after = optimize_one_material(sub, mem_pools.get(m, []), length_col=length_col)
        assign_m["material"] = m

        # Assign global group numbers sequentially in the chosen optimization order.
        # This keeps all Saw #13 groups at the end.
        if not assign_m.empty and "mat_group" in assign_m.columns:
            n_groups = int(assign_m["mat_group"].max())
            assign_m["optimized_group"] = assign_m["mat_group"].astype(int) + (next_global_group - 1)
            next_global_group += n_groups

        parts.append(assign_m)
        updated_pools[m] = pool_after

    opt = pd.concat(parts, ignore_index=True)
    if "mat_group" in opt.columns:
        opt = opt.drop(columns=["mat_group"])

    # Merge back, keep original fields intact
    out = df.merge(opt, on=["orig_index", "material"], how="left")
    out = out.sort_values(["optimized_group", "orig_index"]).reset_index(drop=True)

    # Prepare export view (hide internal/helper/noise columns)
    export_cols_to_drop = {"orig_index", "group"}
    unnamed_cols = {c for c in out.columns if str(c).strip().lower().startswith("unnamed")}
    out_export = out.drop(columns=[c for c in out.columns if c in export_cols_to_drop or c in unnamed_cols], errors="ignore")

    # Build checks (split: main vs TigerStop vs Saw #13) to avoid confusion
    check_all = (
        out.groupby("optimized_group")
        .agg(
            capacity=("source_length_mm", "first"),
            sum_lengths=("length", "sum"),
            material=("material", "first"),
            materials=("material", lambda s: ",".join(sorted(set(str(x) for x in s if pd.notna(x))))),
        )
        .reset_index()
    )
    check_all["ok_capacity"] = check_all["sum_lengths"] <= check_all["capacity"] + 1e-9
    check_all["ok_single_material"] = ~check_all["materials"].str.contains(",")
    check_all["is_saw13"] = check_all["material"].apply(is_saw13_material)
    check_all["is_tigerstop"] = check_all["material"].apply(is_tigerstop_material)

    check_all["is_main_saw"] = check_all["material"].apply(is_main_saw_material)

    check_main = check_all.loc[check_all["is_main_saw"]].drop(columns=["is_saw13", "is_tigerstop", "is_main_saw"]).reset_index(drop=True)
    check_tigerstop = check_all.loc[check_all["is_tigerstop"]].drop(columns=["is_saw13", "is_tigerstop", "is_main_saw"]).reset_index(drop=True)
    check_saw13 = check_all.loc[check_all["is_saw13"]].drop(columns=["is_saw13", "is_tigerstop", "is_main_saw"]).reset_index(drop=True)

    saw13_minimum_sheet = pd.DataFrame(columns=out_export.columns)
    saw13_minimum_mask = pd.Series(False, index=out_export.index)
    if not out_export.empty:
        saw13_minimum_mask = (
            out_export["material"].apply(is_saw13_material)
            & pd.to_numeric(out_export["length"], errors="coerce").lt(MIN_SAW13_CUT_LENGTH)
        )
        saw13_minimum_sheet = out_export.loc[saw13_minimum_mask].copy().reset_index(drop=True)

    # --- Update offcut memory for NEXT batches ---
    # IMPORTANT: Offcuts for materials NOT in this batch must still carry forward.
    batch_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    timestamp = datetime.now().isoformat(timespec="seconds")

    new_mem_records = []

    # 1) Carry forward uploaded offcuts for materials not present in this cut batch
    if not offcut_mem_df.empty:
        carryover_other = offcut_mem_df.loc[~offcut_mem_df["material"].isin(materials_ordered)].copy()
        carryover_other = carryover_other[carryover_other["offcut_length"] > MIN_REUSE]
        for _, r in carryover_other.iterrows():
            new_mem_records.append({
                "material": r["material"],
                "offcut_length": float(r["offcut_length"]),
                "precut_mm": float(r.get("precut_mm", precut_mm)),
                "batch_id": batch_id,
                "timestamp": timestamp,
            })

    # 2) Add remaining pools (unused + newly generated offcuts) for materials in this batch
    for m, pool in updated_pools.items():
        for oc in pool:
            if oc > MIN_REUSE:
                new_mem_records.append({
                    "material": m,
                    "offcut_length": oc,
                    "precut_mm": precut_mm,
                    "batch_id": batch_id,
                    "timestamp": timestamp,
                })

    # 2) Create DataFrame (no longer saving to file for cloud hosting)
    new_mem_df = pd.DataFrame(new_mem_records, columns=["material", "offcut_length", "precut_mm", "batch_id", "timestamp"])

    # 3) Split export into 3 separate files
    main_mask = out_export["material"].apply(is_main_saw_material)
    saw13_mask = out_export["material"].apply(is_saw13_material)
    tiger_mask = out_export["material"].apply(is_tigerstop_material)
    out_export_main = out_export.loc[main_mask].reset_index(drop=True)
    out_export_tigerstop = out_export.loc[tiger_mask].reset_index(drop=True)
    out_export_saw13 = out_export.loc[saw13_mask].reset_index(drop=True)

    # Keep all Saw #13 rows in the master sheet; flag sub-minimum rows for manual cut.
    if not out_export_saw13.empty:
        _saw13_len_num = pd.to_numeric(out_export_saw13["length"], errors="coerce")
        _is_min = _saw13_len_num.lt(MIN_SAW13_CUT_LENGTH)
        out_export_saw13["manual_cut_required"] = _is_min
        out_export_saw13["manual_cut_note"] = ""
        out_export_saw13.loc[_is_min, "manual_cut_note"] = (
            f"MIN SIZE (<{int(MIN_SAW13_CUT_LENGTH)}mm) - MACHINE CAN'T CUT; MANUAL CUT REQUIRED"
        )

    main_excel_buf = io.BytesIO()
    with pd.ExcelWriter(main_excel_buf, engine="openpyxl") as writer:
        out_export_main.to_excel(writer, sheet_name="main saw", index=False)
        check_main.to_excel(writer, sheet_name="Checks", index=False)
    main_excel_buf.seek(0)

    tigerstop_excel_buf = io.BytesIO()
    with pd.ExcelWriter(tigerstop_excel_buf, engine="openpyxl") as writer:
        out_export_tigerstop.to_excel(writer, sheet_name="Tiger Stop Saw", index=False)
        check_tigerstop.to_excel(writer, sheet_name="Checks", index=False)
    tigerstop_excel_buf.seek(0)

    # For material 5956.05.00.4880 only, force qty to 1 in Saw #13 sheet
    _qty_col_saw13 = next((c for c in out_export_saw13.columns if str(c).strip().lower() in ("qty", "quantity")), None)
    if _qty_col_saw13 is not None:
        _mask_5956 = out_export_saw13["material"].astype(str).str.strip() == "5956.05.00.4880"
        out_export_saw13.loc[_mask_5956, _qty_col_saw13] = 1

    saw13_excel_buf = io.BytesIO()
    with pd.ExcelWriter(saw13_excel_buf, engine="openpyxl") as writer:
        out_export_saw13.to_excel(writer, sheet_name="saw #13", index=False)
        check_saw13.to_excel(writer, sheet_name="Checks", index=False)
        if not saw13_minimum_sheet.empty:
            saw13_minimum_sheet.to_excel(writer, sheet_name="minimum size saw #13", index=False)

        # Visually highlight rows that require manual cutting (< minimum machine cut length).
        if not out_export_saw13.empty and "manual_cut_required" in out_export_saw13.columns:
            from openpyxl.styles import PatternFill

            ws = writer.sheets.get("saw #13")
            if ws is not None:
                fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")
                manual_idx = out_export_saw13.index[out_export_saw13["manual_cut_required"]].tolist()
                ncols = len(out_export_saw13.columns)
                for ridx in manual_idx:
                    excel_row = int(ridx) + 2  # +1 header, +1 1-based index
                    for c in range(1, ncols + 1):
                        ws.cell(row=excel_row, column=c).fill = fill
    saw13_excel_buf.seek(0)

    return out, check_main, check_tigerstop, check_saw13, main_excel_buf, tigerstop_excel_buf, saw13_excel_buf, new_mem_df

# ---------- Run when file uploaded ----------
if uploaded is not None:
    st.success("✅ File uploaded. Click to generate optimized output using saved offcuts first.")
    if st.button("Generate optimized files"):
        try:
            if uploaded.name.lower().endswith(".csv"):
                df_input = pd.read_csv(uploaded, sep=None, engine="python")  # auto-detect delimiter
            else:
                df_input = pd.read_excel(uploaded, engine="openpyxl")
        except Exception as e:
            st.error(f"❌ Could not read the file: {e}")
            st.stop()
        
        # Use uploaded offcuts if provided, otherwise optimize without offcuts
        if uploaded_offcuts is not None:
            try:
                offcut_to_use = pd.read_csv(uploaded_offcuts)
                st.info(f"ℹ️ Using uploaded offcut inventory ({len(offcut_to_use)} offcuts)")
            except Exception as e:
                st.warning(f"⚠️ Could not read uploaded offcut file: {e}. Optimizing without offcuts.")
                offcut_to_use = pd.DataFrame(columns=["material", "offcut_length", "precut_mm", "batch_id", "timestamp"])
        else:
            offcut_to_use = pd.DataFrame(columns=["material", "offcut_length", "precut_mm", "batch_id", "timestamp"])
            st.info("ℹ️ No offcut inventory uploaded. Optimizing with fresh stock only.")

        out, check_main, check_tigerstop, check_saw13, main_excel_buf, tigerstop_excel_buf, saw13_excel_buf, new_mem_df = optimize_with_memory(df_input, offcut_to_use, precut_mm)

        # Create machine-readable file ONLY for main saw (exclude Saw #13 materials)
        try:
            out_main_for_machine = out.loc[out["material"].apply(is_main_saw_material)].copy()
            machine_df = transform_optimized_to_machine_readable(out_main_for_machine)
            machine_buf = io.BytesIO()
            machine_df.to_csv(machine_buf, index=False, header=False, encoding="utf-8")
            machine_buf.seek(0)
        except Exception as e:
            machine_buf = None
            st.warning(f"⚠️ Could not generate machine readable file: {e}")

        # Create Saw #13 JSON from optimized output (additive; does not affect existing files)
        saw13_json_text = None
        saw13_json_summary = None
        try:
            out_saw13_for_json = out.loc[
                out["material"].apply(is_saw13_material)
                & pd.to_numeric(out["length"], errors="coerce").ge(MIN_SAW13_CUT_LENGTH)
            ].copy()
            if not out_saw13_for_json.empty:
                saw13_json_data, saw13_json_summary = generate_saw13_json_from_optimized(out_saw13_for_json)
                saw13_json_text = json.dumps(saw13_json_data, ensure_ascii=False, indent=2)
        except Exception as e:
            saw13_json_text = None
            saw13_json_summary = None
            st.warning(f"⚠️ Could not generate Saw #13 JSON: {e}")

        # Reorder barcode PDF (if provided) to match optimized sequence
        barcode_buf = None
        barcode_report = None
        if barcode_pdf is not None:
            try:
                pdf_bytes = barcode_pdf.getvalue()
                barcode_buf, barcode_report = reorder_barcode_pdf_to_optimized(pdf_bytes, out)
            except Exception as e:
                barcode_buf = None
                barcode_report = None
                st.warning(f"⚠️ Could not reorder barcode PDF: {e}")

        # Split barcode PDF by machine — use reordered PDF so per-machine PDFs
        # are already in optimized order. Fall back to original if reorder failed.
        split_barcode_pdfs = {}
        if barcode_pdf is not None:
            try:
                split_source = barcode_buf.getvalue() if barcode_buf is not None else barcode_pdf.getvalue()
                split_barcode_pdfs = split_barcode_pdf_by_machine(split_source)
            except Exception as e:
                st.warning(f"⚠️ Could not split barcode PDF by machine: {e}")

        # Store results in session state
        st.session_state.results = {
            'main_excel_buf': main_excel_buf,
            'tigerstop_excel_buf': tigerstop_excel_buf,
            'saw13_excel_buf': saw13_excel_buf,
            'check_main': check_main,
            'check_tigerstop': check_tigerstop,
            'check_saw13': check_saw13,
            'new_mem_df': new_mem_df,
            'machine_buf': machine_buf,
            'saw13_json_text': saw13_json_text,
            'saw13_json_summary': saw13_json_summary,
            'barcode_buf': barcode_buf,
            'barcode_report': barcode_report,
            'split_barcode_main': split_barcode_pdfs.get('main_saw'),
            'split_barcode_saw13': split_barcode_pdfs.get('saw13'),
            'split_barcode_tigerstop': split_barcode_pdfs.get('tigerstop'),
        }
    
    # Display downloads if results exist in session state
    if 'results' in st.session_state:
        st.subheader("📂 Downloads")
        st.download_button(
            label="⬇️ Main saw Excel (.xlsx)",
            data=st.session_state.results['main_excel_buf'],
            file_name="main saw.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.download_button(
            label="⬇️ Tiger Stop Saw Excel (.xlsx)",
            data=st.session_state.results['tigerstop_excel_buf'],
            file_name="Tiger Stop Saw.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        st.download_button(
            label="⬇️ Saw #13 Excel (.xlsx)",
            data=st.session_state.results['saw13_excel_buf'],
            file_name="saw #13.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if st.session_state.results.get('machine_buf') is not None:
            st.download_button(
                label="⬇️ Main saw machine readable format",
                data=st.session_state.results['machine_buf'],
                file_name="machine_readable_profile.csv",
                mime="text/csv",
            )

        if st.session_state.results.get('saw13_json_text') is not None:
            st.download_button(
                label="⬇️ Saw #13 JSON",
                data=st.session_state.results['saw13_json_text'].encode('utf-8'),
                file_name="saw13_generated_working.json",
                mime="application/json",
            )

        # Split barcode PDF downloads
        if st.session_state.results.get('split_barcode_main') is not None:
            st.download_button(
                label="⬇️ Barcode labels - Main Saw",
                data=st.session_state.results['split_barcode_main'],
                file_name="barcodes_main_saw.pdf",
                mime="application/pdf",
            )
        if st.session_state.results.get('split_barcode_saw13') is not None:
            st.download_button(
                label="⬇️ Barcode labels - Saw #13",
                data=st.session_state.results['split_barcode_saw13'],
                file_name="barcodes_saw13.pdf",
                mime="application/pdf",
            )
        if st.session_state.results.get('split_barcode_tigerstop') is not None:
            st.download_button(
                label="⬇️ Barcode labels - Tiger Stop",
                data=st.session_state.results['split_barcode_tigerstop'],
                file_name="barcodes_tigerstop.pdf",
                mime="application/pdf",
            )

        if st.session_state.results.get('barcode_report') is not None:
            st.caption("Barcode reorder summary")
            st.dataframe(st.session_state.results['barcode_report'], use_container_width=True)

        if st.session_state.results.get('saw13_json_summary') is not None:
            saw13_summary = st.session_state.results['saw13_json_summary']
            if saw13_summary.get('missing_maxy_materials_defaulted_to_45'):
                st.warning("Saw #13 JSON: Some materials were not in MaxY map; default 45.0 was used.")
                st.write(saw13_summary.get('missing_maxy_materials_defaulted_to_45'))
            if saw13_summary.get('over_length_groups'):
                st.warning(
                    "Saw #13 JSON: Some groups exceed source_length_mm after qty expansion. "
                    "BarLeftLength was written as 0 for those groups."
                )
                st.dataframe(pd.DataFrame(saw13_summary.get('over_length_groups')), use_container_width=True)
        
        # Download button for generated offcuts
        if not st.session_state.results['new_mem_df'].empty:
            offcut_download = st.session_state.results['new_mem_df'][["material", "offcut_length", "precut_mm"]].to_csv(index=False).encode('utf-8')
            st.download_button(
                label="⬇️ Download Generated Offcuts Inventory",
                data=offcut_download,
                file_name="reusable_offcuts_inventory.csv",
                mime="text/csv",
                help="Save this file and upload it next time to reuse these offcuts"
            )

        st.subheader("🧧 Generated Reusable Offcuts")
        if not st.session_state.results['new_mem_df'].empty:
            st.dataframe(
                st.session_state.results['new_mem_df'][["material", "offcut_length"]].sort_values(["material", "offcut_length"], ascending=[True, False]),
                use_container_width=True
            )
            st.caption(f"💾 Save the offcut inventory above and upload it next time to reuse {len(st.session_state.results['new_mem_df'])} offcuts")
        else:
            st.info("No reusable offcuts generated (all cuts used full stock or waste < 330mm)")
else:
    st.info("Upload an Excel or CSV file to begin.")
